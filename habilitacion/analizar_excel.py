"""
Script para analizar el Excel de autoevaluación y entender la estructura
de títulos, subtítulos y criterios evaluables.
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'habilitacion_project.settings')
django.setup()

import openpyxl
from openpyxl.styles import Font

# Cargar el archivo Excel
wb = openpyxl.load_workbook('autoevaluacion-_resolucion-3100-2019-anexo-estandar.xlsx')

print("=" * 80)
print("ANÁLISIS DEL EXCEL DE AUTOEVALUACIÓN")
print("=" * 80)

# Listar todas las hojas
print(f"\nTotal de hojas: {len(wb.sheetnames)}")
print("\nHojas disponibles:")
for i, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"  {i}. {sheet_name}")

# Analizar la primera hoja en detalle (11.1.1TH - Talento Humano)
print("\n" + "=" * 80)
print("ANÁLISIS DETALLADO DE LA HOJA '11.1.1TH'")
print("=" * 80)

sheet = wb['11.1.1TH']

# Ver las primeras filas para entender estructura
print("\nPrimeras 30 filas:")
print("-" * 80)

for row_num in range(1, 31):
    row_data = []
    is_bold = False
    is_merged = False

    for col in range(1, 5):  # Columnas A-D
        cell = sheet.cell(row=row_num, column=col)
        value = cell.value

        # Verificar si es negrita
        if cell.font and cell.font.bold:
            is_bold = True

        # Verificar si está en celdas fusionadas
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                is_merged = True
                break

        if value:
            row_data.append(str(value)[:50])
        else:
            row_data.append("")

    # Solo mostrar filas con contenido
    if any(row_data):
        tipo = ""
        if is_bold and is_merged:
            tipo = "[TÍTULO PRINCIPAL]"
        elif is_bold:
            tipo = "[SUBTÍTULO/SECCIÓN]"
        else:
            tipo = "[CRITERIO]"

        print(f"Fila {row_num:3d} {tipo:20s}: {' | '.join(row_data)}")

# Analizar patrones en todas las hojas
print("\n" + "=" * 80)
print("PATRONES IDENTIFICADOS EN TODAS LAS HOJAS")
print("=" * 80)

for sheet_name in wb.sheetnames[:5]:  # Analizar primeras 5 hojas
    sheet = wb[sheet_name]
    print(f"\n--- {sheet_name} ---")

    titulos = 0
    criterios = 0

    for row_num in range(1, sheet.max_row + 1):
        cell_a = sheet.cell(row=row_num, column=1)
        cell_b = sheet.cell(row=row_num, column=2)

        if cell_a.value or cell_b.value:
            # Si columna A tiene valor y B está vacío, probablemente es título
            # Si columna A tiene "Estándar" o número, y B tiene texto largo, es criterio

            is_bold_a = cell_a.font and cell_a.font.bold if cell_a.font else False
            is_bold_b = cell_b.font and cell_b.font.bold if cell_b.font else False

            col_a = str(cell_a.value or "").strip()
            col_b = str(cell_b.value or "").strip()

            # Detectar títulos (generalmente en negrita y celdas fusionadas o solo col A)
            if is_bold_a and (not col_b or is_bold_b):
                titulos += 1
                if row_num <= 20:  # Mostrar primeros títulos
                    print(f"  TÍTULO fila {row_num}: {col_a[:60]}")
            elif col_b and len(col_b) > 20:  # Criterios tienen texto largo en col B
                criterios += 1

    print(f"  Total títulos/secciones: {titulos}")
    print(f"  Total criterios: {criterios}")

# Analizar columnas y encabezados
print("\n" + "=" * 80)
print("ESTRUCTURA DE COLUMNAS")
print("=" * 80)

sheet = wb['11.1.1TH']
print("\nEncabezados detectados:")
for col in range(1, 10):
    cell = sheet.cell(row=1, column=col)
    if cell.value:
        print(f"  Columna {col}: {cell.value}")

# Buscar la fila de encabezados real
for row_num in range(1, 10):
    row_values = [sheet.cell(row=row_num, column=col).value for col in range(1, 5)]
    if any(val and 'Estándar' in str(val) for val in row_values if val):
        print(f"\nFila de encabezados encontrada en fila {row_num}:")
        for col, val in enumerate(row_values, 1):
            if val:
                print(f"  Col {col}: {val}")
        break

print("\n" + "=" * 80)
print("FIN DEL ANÁLISIS")
print("=" * 80)
