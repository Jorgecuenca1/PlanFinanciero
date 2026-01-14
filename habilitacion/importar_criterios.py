"""
Script para importar criterios desde el Excel de la Resolucion 3100 de 2019
"""
import os
import sys
import django

# Configurar Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'habilitacion_project.settings')
django.setup()

import openpyxl
from estandares.models import GrupoEstandar, Estandar, Criterio

# Mapeo de hojas a estandares
HOJAS_ESTANDARES = {
    '11.1.1TH': ('11.1', '11.1.1', 'Talento Humano'),
    '11.1.2.INF': ('11.1', '11.1.2', 'Infraestructura'),
    '11.1.3.DOT': ('11.1', '11.1.3', 'Dotacion'),
    '11.1.4MD': ('11.1', '11.1.4', 'Medicamentos, Dispositivos Medicos e Insumos'),
    '11.1.5.PP': ('11.1', '11.1.5', 'Procesos Prioritarios'),
    '11.1.6.HCR': ('11.1', '11.1.6', 'Historia Clinica y Registros'),
    '11.1.7INT': ('11.1', '11.1.7', 'Interdependencia'),
}

def importar_criterios():
    """Importa todos los criterios desde el Excel"""

    # Cargar el Excel
    print("Cargando archivo Excel...")
    wb = openpyxl.load_workbook('autoevaluacion-_resolucion-3100-2019-anexo-estandar.xlsx', read_only=True)

    total_criterios = 0

    # Asegurar que existe el grupo 11.1
    grupo, _ = GrupoEstandar.objects.get_or_create(
        codigo='11.1',
        defaults={
            'nombre': 'Estandares aplicables a todos los servicios',
            'descripcion': 'Estandares que aplican a todos los servicios de salud',
            'aplica_todos': True,
            'orden': 1,
            'activo': True
        }
    )
    print(f"Grupo: {grupo.codigo} - {grupo.nombre}")

    # Procesar cada hoja
    for hoja_nombre, (grupo_codigo, estandar_codigo, estandar_nombre) in HOJAS_ESTANDARES.items():
        print(f"\nProcesando hoja: {hoja_nombre}")

        if hoja_nombre not in wb.sheetnames:
            print(f"  ADVERTENCIA: Hoja '{hoja_nombre}' no encontrada")
            continue

        # Obtener o crear el estandar
        estandar, created = Estandar.objects.get_or_create(
            codigo=estandar_codigo,
            grupo=grupo,
            defaults={
                'nombre': estandar_nombre,
                'descripcion': f'Estandar de {estandar_nombre}',
                'orden': int(estandar_codigo.split('.')[-1]),
                'activo': True
            }
        )

        if created:
            print(f"  Estandar creado: {estandar.codigo} - {estandar.nombre}")
        else:
            print(f"  Estandar existente: {estandar.codigo} - {estandar.nombre}")

        # Eliminar criterios existentes para reimportar
        Criterio.objects.filter(estandar=estandar).delete()

        ws = wb[hoja_nombre]
        criterios_hoja = 0
        orden = 0

        for row in ws.iter_rows(min_row=4, values_only=True):  # Empezar desde fila 4
            codigo_std = row[0] if row[0] else ''
            texto = row[1] if row[1] else ''

            # Saltar filas vacias o sin criterio
            if not texto or not texto.strip():
                continue

            texto = texto.strip()

            # Determinar si es titulo (sin numero al inicio)
            es_titulo = False
            numero = ''

            # Extraer numero del criterio
            if texto[0].isdigit():
                # Buscar el punto o espacio despues del numero
                for i, char in enumerate(texto):
                    if char == '.' or char == ' ':
                        numero = texto[:i+1].strip().rstrip('.')
                        break
                    elif not char.isdigit() and char != '.':
                        break

            if not numero:
                # Podria ser un titulo o un criterio sin numero
                es_titulo = ':' in texto[:50] or texto.endswith(':')

            orden += 1

            # Crear el criterio
            criterio = Criterio.objects.create(
                estandar=estandar,
                numero=numero or str(orden),
                texto=texto,
                es_titulo=es_titulo,
                orden=orden,
                activo=True
            )
            criterios_hoja += 1

        print(f"  Criterios importados: {criterios_hoja}")
        total_criterios += criterios_hoja

    print(f"\n{'='*50}")
    print(f"TOTAL CRITERIOS IMPORTADOS: {total_criterios}")
    print(f"{'='*50}")

    wb.close()

if __name__ == '__main__':
    importar_criterios()
