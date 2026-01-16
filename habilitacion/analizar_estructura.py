"""
Script para analizar mejor la estructura de los criterios en el Excel.
"""
import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'habilitacion_project.settings')
django.setup()

import openpyxl

wb = openpyxl.load_workbook('autoevaluacion-_resolucion-3100-2019-anexo-estandar.xlsx')

# Analizar hoja 11.1.1TH en detalle
sheet = wb['11.1.1TH']

print("ANÁLISIS DETALLADO DE FILAS:")
print("=" * 100)

for row_num in range(1, 35):
    col_a = sheet.cell(row=row_num, column=1).value
    col_b = sheet.cell(row=row_num, column=2).value
    col_c = sheet.cell(row=row_num, column=3).value
    col_d = sheet.cell(row=row_num, column=4).value

    # Verificar si la celda tiene contenido
    if col_a or col_b:
        estado_str = str(col_c).strip() if col_c else ''
        tiene_estado = estado_str.upper() in ['C', 'NC', 'NA']

        print(f"Fila {row_num:2d}:")
        print(f"  Col A: {str(col_a)[:30] if col_a else '(vacío)'}")
        print(f"  Col B: {str(col_b)[:60] if col_b else '(vacío)'}...")
        print(f"  Col C: '{estado_str}' -> tiene_estado={tiene_estado}")
        print()

# Analizar otra hoja (11.1.2.INF) que tiene más contenido
print("\n\nANÁLISIS DE HOJA 11.1.2.INF:")
print("=" * 100)

sheet = wb['11.1.2.INF']

for row_num in range(1, 30):
    col_a = sheet.cell(row=row_num, column=1).value
    col_b = sheet.cell(row=row_num, column=2).value
    col_c = sheet.cell(row=row_num, column=3).value

    if col_a or col_b:
        estado_str = str(col_c).strip() if col_c else ''
        tiene_estado = estado_str.upper() in ['C', 'NC', 'NA']

        # Verificar negrita
        is_bold_a = sheet.cell(row=row_num, column=1).font.bold if sheet.cell(row=row_num, column=1).font else False
        is_bold_b = sheet.cell(row=row_num, column=2).font.bold if sheet.cell(row=row_num, column=2).font else False

        print(f"Fila {row_num:2d}: bold_a={is_bold_a}, bold_b={is_bold_b}, estado='{estado_str}'")
        print(f"  A: {str(col_a)[:30] if col_a else '(vacío)'}")
        print(f"  B: {str(col_b)[:80] if col_b else '(vacío)'}")
        print()
