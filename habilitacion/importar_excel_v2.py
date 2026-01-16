"""
Script MEJORADO para importar los criterios del Excel de autoevaluación.
Identifica correctamente títulos, subtítulos y criterios evaluables.

Lógica de clasificación:
- TITULO: Texto en negrita sin número, o encabezados de sección
- SUBTITULO: Texto que termina en ":" o introducción a un grupo de criterios
- CRITERIO: Texto con número que es evaluable (tiene o puede tener C/NC/NA)
"""
import os
import re
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'habilitacion_project.settings')
django.setup()

import openpyxl
from estandares.models import GrupoEstandar, Estandar, Criterio

# Mapeo de hojas del Excel a estructura
MAPEO_HOJAS = {
    # Grupo 11.1 - Todos los Servicios
    '11.1.1TH': {'grupo': '11.1', 'estandar': '11.1.1', 'nombre': 'Talento Humano', 'codigo_corto': 'TSTH'},
    '11.1.2.INF': {'grupo': '11.1', 'estandar': '11.1.2', 'nombre': 'Infraestructura', 'codigo_corto': 'TSINF'},
    '11.1.3.DOT': {'grupo': '11.1', 'estandar': '11.1.3', 'nombre': 'Dotación', 'codigo_corto': 'TSDOT'},
    '11.1.4MD': {'grupo': '11.1', 'estandar': '11.1.4', 'nombre': 'Medicamentos, Dispositivos Médicos e Insumos', 'codigo_corto': 'TSMD'},
    '11.1.5.PP': {'grupo': '11.1', 'estandar': '11.1.5', 'nombre': 'Procesos Prioritarios', 'codigo_corto': 'TSPP'},
    '11.1.6.HCR': {'grupo': '11.1', 'estandar': '11.1.6', 'nombre': 'Historia Clínica y Registros', 'codigo_corto': 'TSHCR'},
    '11.1.7INT': {'grupo': '11.1', 'estandar': '11.1.7', 'nombre': 'Interdependencia', 'codigo_corto': 'TSINT'},

    # Grupo 11.2 - Consulta Externa
    '11.2.1.S_CE_G': {'grupo': '11.2', 'estandar': '11.2.1', 'nombre': 'Consulta Externa General', 'codigo_corto': 'CEG'},
    '11.2.2.S_CE_E': {'grupo': '11.2', 'estandar': '11.2.2', 'nombre': 'Consulta Externa Especializada', 'codigo_corto': 'CEE'},
    '11.2.3.S_CE_V': {'grupo': '11.2', 'estandar': '11.2.3', 'nombre': 'Consulta Externa Vacunación', 'codigo_corto': 'CEV'},
    '11.2.4.S_CE_SST': {'grupo': '11.2', 'estandar': '11.2.4', 'nombre': 'Consulta Externa SST', 'codigo_corto': 'CESST'},

    # Grupo 11.3 - Apoyo Diagnóstico
    '11.3.1.S_TR': {'grupo': '11.3', 'estandar': '11.3.1', 'nombre': 'Toma de Muestras', 'codigo_corto': 'TR'},
    ' 11.3.2.SF': {'grupo': '11.3', 'estandar': '11.3.2', 'nombre': 'Servicio Farmacéutico', 'codigo_corto': 'SF'},
    '11.3.3.S_Rx_OD': {'grupo': '11.3', 'estandar': '11.3.3', 'nombre': 'Radiología e Imágenes Diagnósticas', 'codigo_corto': 'RxOD'},
    '11.3.4.S_IDx': {'grupo': '11.3', 'estandar': '11.3.4', 'nombre': 'Imágenes Diagnósticas', 'codigo_corto': 'IDx'},
    ' 11.3.5.S_MNuc': {'grupo': '11.3', 'estandar': '11.3.5', 'nombre': 'Medicina Nuclear', 'codigo_corto': 'MNuc'},
    '11.3.6.S_Radio': {'grupo': '11.3', 'estandar': '11.3.6', 'nombre': 'Radioterapia', 'codigo_corto': 'Radio'},
    '11.3.7.S_Quimio': {'grupo': '11.3', 'estandar': '11.3.7', 'nombre': 'Quimioterapia', 'codigo_corto': 'Quimio'},
    '11.3.8 S_Dx VASC': {'grupo': '11.3', 'estandar': '11.3.8', 'nombre': 'Diagnóstico Vascular', 'codigo_corto': 'DxVASC'},
    '11.3.9S_HI': {'grupo': '11.3', 'estandar': '11.3.9', 'nombre': 'Hemodinamia e Intervencionismo', 'codigo_corto': 'HI'},
    '11.3.10 S_GPT': {'grupo': '11.3', 'estandar': '11.3.10', 'nombre': 'Gestión Pre-Transfusional', 'codigo_corto': 'GPT'},
    '11.3.11.S_TMLC': {'grupo': '11.3', 'estandar': '11.3.11', 'nombre': 'Toma Muestras Lab Clínico', 'codigo_corto': 'TMLC'},
    '11.3.12.S_LC': {'grupo': '11.3', 'estandar': '11.3.12', 'nombre': 'Laboratorio Clínico', 'codigo_corto': 'LC'},
    ' 11.3.13 S_TM_CU': {'grupo': '11.3', 'estandar': '11.3.13', 'nombre': 'Toma Muestras Citología', 'codigo_corto': 'TMCU'},
    '11.3.14.S_LCU': {'grupo': '11.3', 'estandar': '11.3.14', 'nombre': 'Laboratorio Citología', 'codigo_corto': 'LCU'},
    '11.3.15.S_LHT': {'grupo': '11.3', 'estandar': '11.3.15', 'nombre': 'Laboratorio Histotecnología', 'codigo_corto': 'LHT'},
    '11.3.16.S_LPT': {'grupo': '11.3', 'estandar': '11.3.16', 'nombre': 'Laboratorio Patología', 'codigo_corto': 'LPT'},
    '11.3.17.S_Dial': {'grupo': '11.3', 'estandar': '11.3.17', 'nombre': 'Diálisis', 'codigo_corto': 'Dial'},

    # Grupo 11.4 - Internación
    '11.4.1.S_HP': {'grupo': '11.4', 'estandar': '11.4.1', 'nombre': 'Hospitalización', 'codigo_corto': 'HP'},
    '11.4.2.S_HP_PC': {'grupo': '11.4', 'estandar': '11.4.2', 'nombre': 'Hospitalización Parcial', 'codigo_corto': 'HPPC'},
    '11.4.3.S_CBN': {'grupo': '11.4', 'estandar': '11.4.3', 'nombre': 'Cuidado Básico Neonatal', 'codigo_corto': 'CBN'},
    '11.4.4.S_CIN ': {'grupo': '11.4', 'estandar': '11.4.4', 'nombre': 'Cuidado Intermedio Neonatal', 'codigo_corto': 'CIN'},
    '11.4.5.S_CINN': {'grupo': '11.4', 'estandar': '11.4.5', 'nombre': 'Cuidado Intensivo Neonatal', 'codigo_corto': 'CINN'},
    '11.4.6. S_CINP': {'grupo': '11.4', 'estandar': '11.4.6', 'nombre': 'Cuidado Intermedio Pediátrico', 'codigo_corto': 'CINP'},
    '11.4.7.S_CIP': {'grupo': '11.4', 'estandar': '11.4.7', 'nombre': 'Cuidado Intensivo Pediátrico', 'codigo_corto': 'CIP'},
    '11.4.8.S_CIMA': {'grupo': '11.4', 'estandar': '11.4.8', 'nombre': 'Cuidado Intermedio Adulto', 'codigo_corto': 'CIMA'},
    '11.4.9.S_CIA': {'grupo': '11.4', 'estandar': '11.4.9', 'nombre': 'Cuidado Intensivo Adulto', 'codigo_corto': 'CIA'},
    '11.4.10.S_HSM CSP': {'grupo': '11.4', 'estandar': '11.4.10', 'nombre': 'Hospitalización Salud Mental', 'codigo_corto': 'HSM'},
    '11.4.11.S_HSP': {'grupo': '11.4', 'estandar': '11.4.11', 'nombre': 'Hospitalización Psiquiátrica', 'codigo_corto': 'HSP'},
    '11.4.12.S_CB_CSP': {'grupo': '11.4', 'estandar': '11.4.12', 'nombre': 'Cuidados Básicos Especiales', 'codigo_corto': 'CBCSP'},

    # Grupo 11.5 - Quirúrgico
    '11.5.1.S_CX': {'grupo': '11.5', 'estandar': '11.5.1', 'nombre': 'Cirugía', 'codigo_corto': 'CX'},

    # Grupo 11.6 - Atención Inmediata
    '11.6.1.S_UR': {'grupo': '11.6', 'estandar': '11.6.1', 'nombre': 'Urgencias', 'codigo_corto': 'UR'},
    '11.6.2.S_TR_AS': {'grupo': '11.6', 'estandar': '11.6.2', 'nombre': 'Transporte Asistencial', 'codigo_corto': 'TRAS'},
    '11.6.3.S_AT_PH': {'grupo': '11.6', 'estandar': '11.6.3', 'nombre': 'Atención Prehospitalaria', 'codigo_corto': 'ATPH'},
    '11.6.4.S_A.parto': {'grupo': '11.6', 'estandar': '11.6.4', 'nombre': 'Atención del Parto', 'codigo_corto': 'AParto'},
}

GRUPOS_INFO = {
    '11.1': {'nombre': 'Estándares aplicables a todos los servicios', 'aplica_todos': True, 'orden': 1},
    '11.2': {'nombre': 'Grupo Consulta Externa', 'aplica_todos': False, 'orden': 2},
    '11.3': {'nombre': 'Grupo Apoyo Diagnóstico y Complementación Terapéutica', 'aplica_todos': False, 'orden': 3},
    '11.4': {'nombre': 'Grupo Internación', 'aplica_todos': False, 'orden': 4},
    '11.5': {'nombre': 'Grupo Quirúrgico', 'aplica_todos': False, 'orden': 5},
    '11.6': {'nombre': 'Grupo Atención Inmediata', 'aplica_todos': False, 'orden': 6},
}


def tiene_numero_inicio(texto):
    """Verifica si el texto empieza con un número (criterio evaluable)."""
    texto = str(texto).strip()
    # Patrón: número al inicio, puede ser 1. o 1.1. o 1.1.1. etc.
    return bool(re.match(r'^\d+\.?\d*\.?\d*\.?\d*\s*\.?\s+', texto))


def extraer_numero_criterio(texto):
    """Extrae el número del criterio del texto."""
    texto = str(texto).strip()
    match = re.match(r'^(\d+\.?\d*\.?\d*\.?\d*)\s*\.?\s*', texto)
    if match:
        return match.group(1).rstrip('.')
    return ''


def determinar_tipo_criterio(texto, es_bold, tiene_estado):
    """
    Determina el tipo de criterio basándose en:
    - texto: contenido de la celda
    - es_bold: si está en negrita
    - tiene_estado: si tiene C/NC/NA en la columna de estado
    """
    texto = str(texto).strip()

    if not texto:
        return None

    # Ignorar filas de resumen
    texto_upper = texto.upper()
    if texto_upper in ['CUMPLE', 'NO CUMPLE', 'NO APLICA', 'TOTAL', 'ESTADO', 'CRITERIOS']:
        return None
    if texto.startswith('='):  # Fórmulas
        return None
    if 'VOLVER A TABLA' in texto_upper:
        return None

    # Si está en negrita y NO tiene número al inicio -> TITULO
    if es_bold and not tiene_numero_inicio(texto):
        return 'TITULO'

    # Si tiene número al inicio -> es potencialmente evaluable
    if tiene_numero_inicio(texto):
        # Si termina con ":" sin más contenido después, es un subtítulo
        # Ejemplo: "13. Procedimientos bajo sedación:"
        if texto.rstrip().endswith(':'):
            return 'SUBTITULO'
        # Si tiene estado o es un ítem con número, es criterio evaluable
        return 'CRITERIO'

    # Si NO tiene número y termina con ":", es subtítulo
    if texto.rstrip().endswith(':'):
        return 'SUBTITULO'

    # Texto sin número y sin ":" -> título de sección
    return 'TITULO'


def importar_hoja(wb, nombre_hoja, info_hoja, estandar_obj):
    """Importa los criterios de una hoja específica."""
    if nombre_hoja not in wb.sheetnames:
        print(f"  [!] Hoja '{nombre_hoja}' no encontrada")
        return {'criterios': 0, 'titulos': 0, 'subtitulos': 0}

    sheet = wb[nombre_hoja]
    stats = {'criterios': 0, 'titulos': 0, 'subtitulos': 0}
    orden = 0

    # Encontrar fila de inicio de datos (después de encabezados)
    fila_inicio = 4  # Generalmente fila 4

    for row_num in range(fila_inicio, sheet.max_row + 1):
        cell_a = sheet.cell(row=row_num, column=1)
        cell_b = sheet.cell(row=row_num, column=2)
        cell_c = sheet.cell(row=row_num, column=3)

        col_a = cell_a.value
        col_b = cell_b.value
        col_c = cell_c.value

        # Saltar filas vacías
        if not col_b:
            continue

        texto = str(col_b).strip()

        # Verificar si tiene estado válido
        estado = str(col_c).strip().upper() if col_c else ''
        tiene_estado = estado in ['C', 'NC', 'NA']

        # Verificar si está en negrita
        es_bold = cell_b.font.bold if cell_b.font else False

        # Determinar tipo
        tipo = determinar_tipo_criterio(texto, es_bold, tiene_estado)

        if not tipo:
            continue

        # Extraer número
        numero = extraer_numero_criterio(texto)
        if not numero:
            numero = f"T{orden}"  # Para títulos sin número

        # Limpiar texto (quitar número del inicio si lo tiene)
        texto_limpio = texto
        if tiene_numero_inicio(texto):
            texto_limpio = re.sub(r'^(\d+\.?\d*\.?\d*\.?\d*)\s*\.?\s*', '', texto).strip()

        orden += 1

        # Crear criterio
        Criterio.objects.create(
            estandar=estandar_obj,
            numero=numero,
            texto=texto_limpio,
            tipo_criterio=tipo,
            es_titulo=(tipo in ['TITULO', 'SUBTITULO']),
            orden=orden,
            activo=True
        )

        if tipo == 'CRITERIO':
            stats['criterios'] += 1
        elif tipo == 'TITULO':
            stats['titulos'] += 1
        elif tipo == 'SUBTITULO':
            stats['subtitulos'] += 1

    return stats


def main():
    print("=" * 80)
    print("IMPORTACIÓN MEJORADA DEL EXCEL DE AUTOEVALUACIÓN")
    print("=" * 80)

    # Cargar Excel
    print("\n[1] Cargando archivo Excel...")
    wb = openpyxl.load_workbook('autoevaluacion-_resolucion-3100-2019-anexo-estandar.xlsx')
    print(f"    Hojas encontradas: {len(wb.sheetnames)}")

    # Limpiar datos existentes
    print("\n[2] Limpiando datos existentes...")
    criterios_eliminados = Criterio.objects.all().delete()[0]
    estandares_eliminados = Estandar.objects.all().delete()[0]
    grupos_eliminados = GrupoEstandar.objects.all().delete()[0]
    print(f"    Criterios eliminados: {criterios_eliminados}")
    print(f"    Estándares eliminados: {estandares_eliminados}")
    print(f"    Grupos eliminados: {grupos_eliminados}")

    # Crear grupos
    print("\n[3] Creando grupos de estándares...")
    grupos = {}
    for codigo, info in GRUPOS_INFO.items():
        grupo = GrupoEstandar.objects.create(
            codigo=codigo,
            nombre=info['nombre'],
            aplica_todos=info['aplica_todos'],
            orden=info['orden'],
            activo=True
        )
        grupos[codigo] = grupo
        print(f"    + {codigo}: {info['nombre']}")

    # Crear estándares e importar criterios
    print("\n[4] Importando estándares y criterios...")
    totales = {'criterios': 0, 'titulos': 0, 'subtitulos': 0}
    total_estandares = 0

    for nombre_hoja, info in MAPEO_HOJAS.items():
        codigo_grupo = info['grupo']
        codigo_estandar = info['estandar']
        nombre_estandar = info['nombre']
        codigo_corto = info['codigo_corto']

        grupo = grupos.get(codigo_grupo)
        if not grupo:
            print(f"  [!] Grupo {codigo_grupo} no encontrado")
            continue

        # Crear estándar
        estandar, created = Estandar.objects.get_or_create(
            grupo=grupo,
            codigo=codigo_estandar,
            defaults={
                'codigo_corto': codigo_corto,
                'nombre': nombre_estandar,
                'orden': total_estandares,
                'activo': True
            }
        )

        if created:
            total_estandares += 1

        # Importar criterios
        stats = importar_hoja(wb, nombre_hoja, info, estandar)
        totales['criterios'] += stats['criterios']
        totales['titulos'] += stats['titulos']
        totales['subtitulos'] += stats['subtitulos']

        total = stats['criterios'] + stats['titulos'] + stats['subtitulos']
        print(f"    {codigo_estandar} {nombre_estandar}: {stats['criterios']} criterios, {stats['titulos']} títulos, {stats['subtitulos']} subtítulos")

    # Resumen
    print("\n" + "=" * 80)
    print("RESUMEN DE IMPORTACIÓN")
    print("=" * 80)
    print(f"  Grupos creados: {len(grupos)}")
    print(f"  Estándares creados: {total_estandares}")
    print(f"\n  Distribución de registros:")
    print(f"    - Criterios evaluables: {totales['criterios']}")
    print(f"    - Títulos de sección: {totales['titulos']}")
    print(f"    - Subtítulos: {totales['subtitulos']}")
    print(f"    - TOTAL: {totales['criterios'] + totales['titulos'] + totales['subtitulos']}")

    print("\n[OK] Importación completada exitosamente!")


if __name__ == '__main__':
    main()
